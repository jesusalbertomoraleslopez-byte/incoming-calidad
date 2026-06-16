import os
import shutil
import pandas as pd
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Definir rutas en el directorio del cerebro de la conversación
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CARPETAS_DIR = os.path.join(BASE_DIR, "carpetas_electronicas")

# Crear directorios
os.makedirs(BASE_DIR, exist_ok=True)
os.makedirs(CARPETAS_DIR, exist_ok=True)

print("Directorios creados con éxito.")

# 1. Copiar y recortar logotipo para Favicon
logo_original = r"C:\Users\albertol\.gemini\antigravity\brain\a7b81bff-4cc8-47ef-ad0e-79397dd11a77\media__1781539457139.jpg"
logo_destino = os.path.join(BASE_DIR, "logo_sigrama.png")
favicon_destino = os.path.join(BASE_DIR, "favicon.png")

try:
    if os.path.exists(logo_original):
        # Convertir a PNG y copiar
        img = Image.open(logo_original)
        img.save(logo_destino, "PNG")
        print("Logotipo principal guardado como PNG en:", logo_destino)
        
        # El isotipo rojo está a la izquierda. Su forma es aproximadamente cuadrada.
        # Recortamos la sección izquierda (de 0 a la altura de la imagen en ancho y alto)
        width, height = img.size
        crop_width = int(height * 1.05)
        if crop_width > width:
            crop_width = width
        
        isotipo = img.crop((0, 0, crop_width, height))
        isotipo.save(favicon_destino, "PNG")
        print("Favicon (isotipo rojo recortado) guardado en:", favicon_destino)
    else:
        print("⚠️ Advertencia: Logotipo original no encontrado en la ruta especificada.")
except Exception as e:
    print("⚠️ Error procesando logotipo:", e)

# 2. Inicializar base de datos de Parámetros de Materia Prima
db_parametros_path = os.path.join(BASE_DIR, "BD_Parametros_Materia_Prima.xlsx")
if not os.path.exists(db_parametros_path):
    columnas_parametros = [
        "SKU", "Nombre", "Tipo_Lamina", "Grado_Acero",
        "Espesor_Nominal_in", "Espesor_Tolerancia_Min_in", "Espesor_Tolerancia_Max_in",
        "Ancho_Nominal_in", "Ancho_Tolerancia_Min_in", "Ancho_Tolerancia_Max_in",
        "Largo_Nominal_in", "Largo_Tolerancia_Min_in", "Largo_Tolerancia_Max_in",
        "Zinc_Nominal_oz_ft2", "Zinc_Min_oz_ft2", "Aceitado_Requerido", "Dureza_Max_HRB"
    ]
    df_parametros = pd.DataFrame([
        {
            "SKU": "SKU-GALV-10",
            "Nombre": "Lámina Galvanizada Calibre 10",
            "Tipo_Lamina": "Galvanizada",
            "Grado_Acero": "ASTM A653 CS Tipo B",
            "Espesor_Nominal_in": 0.138,
            "Espesor_Tolerancia_Min_in": -0.008,
            "Espesor_Tolerancia_Max_in": 0.008,
            "Ancho_Nominal_in": 48.00,
            "Ancho_Tolerancia_Min_in": -0.125,
            "Ancho_Tolerancia_Max_in": 0.125,
            "Largo_Nominal_in": 120.00,
            "Largo_Tolerancia_Min_in": -0.250,
            "Largo_Tolerancia_Max_in": 0.250,
            "Zinc_Nominal_oz_ft2": 0.60,
            "Zinc_Min_oz_ft2": 0.60,
            "Aceitado_Requerido": "No",
            "Dureza_Max_HRB": 75
        },
        {
            "SKU": "SKU-GALV-12",
            "Nombre": "Lámina Galvanizada Calibre 12",
            "Tipo_Lamina": "Galvanizada",
            "Grado_Acero": "ASTM A653 CS Tipo B",
            "Espesor_Nominal_in": 0.108,
            "Espesor_Tolerancia_Min_in": -0.008,
            "Espesor_Tolerancia_Max_in": 0.008,
            "Ancho_Nominal_in": 48.00,
            "Ancho_Tolerancia_Min_in": -0.125,
            "Ancho_Tolerancia_Max_in": 0.125,
            "Largo_Nominal_in": 120.00,
            "Largo_Tolerancia_Min_in": -0.250,
            "Largo_Tolerancia_Max_in": 0.250,
            "Zinc_Nominal_oz_ft2": 0.60,
            "Zinc_Min_oz_ft2": 0.60,
            "Aceitado_Requerido": "No",
            "Dureza_Max_HRB": 75
        },
        {
            "SKU": "SKU-GALV-14",
            "Nombre": "Lámina Galvanizada Calibre 14",
            "Tipo_Lamina": "Galvanizada",
            "Grado_Acero": "ASTM A653 CS Tipo B",
            "Espesor_Nominal_in": 0.079,
            "Espesor_Tolerancia_Min_in": -0.008,
            "Espesor_Tolerancia_Max_in": 0.008,
            "Ancho_Nominal_in": 48.00,
            "Ancho_Tolerancia_Min_in": -0.120,
            "Ancho_Tolerancia_Max_in": 0.120,
            "Largo_Nominal_in": 120.00,
            "Largo_Tolerancia_Min_in": -0.250,
            "Largo_Tolerancia_Max_in": 0.250,
            "Zinc_Nominal_oz_ft2": 0.60,
            "Zinc_Min_oz_ft2": 0.60,
            "Aceitado_Requerido": "No",
            "Dureza_Max_HRB": 75
        },
        {
            "SKU": "SKU-GALV-16",
            "Nombre": "Lámina Galvanizada Calibre 16",
            "Tipo_Lamina": "Galvanizada",
            "Grado_Acero": "ASTM A653 CS Tipo B",
            "Espesor_Nominal_in": 0.064,
            "Espesor_Tolerancia_Min_in": -0.008,
            "Espesor_Tolerancia_Max_in": 0.008,
            "Ancho_Nominal_in": 48.00,
            "Ancho_Tolerancia_Min_in": -0.120,
            "Ancho_Tolerancia_Max_in": 0.120,
            "Largo_Nominal_in": 120.00,
            "Largo_Tolerancia_Min_in": -0.250,
            "Largo_Tolerancia_Max_in": 0.250,
            "Zinc_Nominal_oz_ft2": 0.60,
            "Zinc_Min_oz_ft2": 0.60,
            "Aceitado_Requerido": "No",
            "Dureza_Max_HRB": 75
        },
        {
            "SKU": "SKU-DECP-12",
            "Nombre": "Lámina Decapada Calibre 12",
            "Tipo_Lamina": "Decapada",
            "Grado_Acero": "SAE 1008 CS",
            "Espesor_Nominal_in": 0.105,
            "Espesor_Tolerancia_Min_in": -0.008,
            "Espesor_Tolerancia_Max_in": 0.008,
            "Ancho_Nominal_in": 48.00,
            "Ancho_Tolerancia_Min_in": -0.125,
            "Ancho_Tolerancia_Max_in": 0.125,
            "Largo_Nominal_in": 120.00,
            "Largo_Tolerancia_Min_in": -0.250,
            "Largo_Tolerancia_Max_in": 0.250,
            "Zinc_Nominal_oz_ft2": 0.0,
            "Zinc_Min_oz_ft2": 0.0,
            "Aceitado_Requerido": "Sí",
            "Dureza_Max_HRB": 65
        },
        {
            "SKU": "SKU-DECP-14",
            "Nombre": "Lámina Decapada Calibre 14",
            "Tipo_Lamina": "Decapada",
            "Grado_Acero": "SAE 1008 CS",
            "Espesor_Nominal_in": 0.075,
            "Espesor_Tolerancia_Min_in": -0.008,
            "Espesor_Tolerancia_Max_in": 0.008,
            "Ancho_Nominal_in": 48.00,
            "Ancho_Tolerancia_Min_in": -0.125,
            "Ancho_Tolerancia_Max_in": 0.125,
            "Largo_Nominal_in": 120.00,
            "Largo_Tolerancia_Min_in": -0.250,
            "Largo_Tolerancia_Max_in": 0.250,
            "Zinc_Nominal_oz_ft2": 0.0,
            "Zinc_Min_oz_ft2": 0.0,
            "Aceitado_Requerido": "Sí",
            "Dureza_Max_HRB": 65
        },
        {
            "SKU": "SKU-DECP-16",
            "Nombre": "Lámina Decapada Calibre 16",
            "Tipo_Lamina": "Decapada",
            "Grado_Acero": "SAE 1008 CS",
            "Espesor_Nominal_in": 0.060,
            "Espesor_Tolerancia_Min_in": -0.008,
            "Espesor_Tolerancia_Max_in": 0.008,
            "Ancho_Nominal_in": 48.00,
            "Ancho_Tolerancia_Min_in": -0.125,
            "Ancho_Tolerancia_Max_in": 0.125,
            "Largo_Nominal_in": 120.00,
            "Largo_Tolerancia_Min_in": -0.250,
            "Largo_Tolerancia_Max_in": 0.250,
            "Zinc_Nominal_oz_ft2": 0.0,
            "Zinc_Min_oz_ft2": 0.0,
            "Aceitado_Requerido": "Sí",
            "Dureza_Max_HRB": 65
        }
    ], columns=columnas_parametros)
    df_parametros.to_excel(db_parametros_path, index=False, sheet_name="Materia_Prima")
    print("Base de datos de parámetros creada.")

# 3. Inicializar base de datos de Reportes Generales (Incoming Logs)
db_reportes_path = os.path.join(BASE_DIR, "BD_Reportes_Incoming.xlsx")
if not os.path.exists(db_reportes_path):
    df_reportes = pd.DataFrame(columns=[
        "Folio", "Fecha", "Hora", "Proveedor", "Orden_Compra", 
        "Factura_Remision", "Inspector", "Estatus_General", "Dossier_Ruta"
    ])
    df_reportes.to_excel(db_reportes_path, index=False, sheet_name="Recepciones")
    print("Base de datos de reportes generales creada.")

# 4. Inicializar base de datos de Detalle de Atados
db_atados_path = os.path.join(BASE_DIR, "BD_Atados_Incoming.xlsx")
if not os.path.exists(db_atados_path):
    df_atados = pd.DataFrame(columns=[
        "ID_Atado", "Folio", "ID_Atado_Proveedor", "SKU", "Grado_Acero", 
        "Num_Colada", "Lote_Heat", 
        "Espesor_Medido_1_in", "Espesor_Medido_2_in", "Espesor_Medido_3_in", 
        "Espesor_Medido_4_in", "Espesor_Medido_5_in", "Espesor_Medido_6_in", 
        "Espesor_Medido_7_in", "Espesor_Medido_8_in", "Espesor_Medido_9_in", 
        "Espesor_Medido_10_in", "Espesor_Medido_11_in", "Espesor_Medido_12_in", 
        "Ancho_Medido_in", "Largo_Medido_in", "Cantidad_Hojas",
        "Peso_Total_Kg", "Peso_Total_Lb", "Zinc_Medido_oz_ft2", 
        "Dureza_Medida_HRB", "Aceitado_OK", "Defectos_Visuales", 
        "Ubicacion_Almacen", "Estatus_Calidad", "Observaciones"
    ])
    df_atados.to_excel(db_atados_path, index=False, sheet_name="Atados_Detalle")
    print("Base de datos de detalle de atados creada.")

# 5. Generar plantilla Excel estilizada corporativa para carga de mediciones
plantilla_path = os.path.join(BASE_DIR, "plantilla_incoming_calidad.xlsx")
# Siempre recreamos la plantilla para asegurar que tenga la estructura más reciente
wb = Workbook()
ws = wb.active
ws.title = "Incoming_Mediciones"

# Fila 1: Títulos de Hojas agrupadas
ws.cell(row=1, column=4, value="HOJA 1")
ws.merge_cells('D1:F1')
ws.cell(row=1, column=7, value="HOJA 2")
ws.merge_cells('G1:I1')
ws.cell(row=1, column=10, value="HOJA 3")
ws.merge_cells('J1:L1')
ws.cell(row=1, column=13, value="HOJA 4")
ws.merge_cells('M1:O1')

headers = [
    "No_Atado", "Calibre", "Galvanizado_o_Decapado", 
    "Espesor_1", "Espesor_2", "Espesor_3", 
    "Espesor_4", "Espesor_5", "Espesor_6", 
    "Espesor_7", "Espesor_8", "Espesor_9", 
    "Espesor_4", "Espesor_5", "Espesor_6", 
    "Cantidad_Hojas", "Peso_Total_Kg", 
    "Num_Colada", "Lote_Heat", 
    "Observaciones"
]

# Escribir fila 2 (Encabezados)
for col_idx, h in enumerate(headers, 1):
    ws.cell(row=2, column=col_idx, value=h)

# Filas ejemplo descriptivas (Fila 3 y Fila 4)
example_row = [
    "ACEROMEX 01", "CAL 16", "Decapado",
    0.059, 0.059, 0.060, 0.059, 0.059, 0.059, 0.059, 0.060, 0.059, 0.059, 0.060, 0.060,
    45, 2029,
    "COL-2603210220", "LOT-23587",
    "Material con daños en el acabado (óxido)"
]
ws.append(example_row)

example_row_dec = [
    "ACEROMEX 02", "CAL 16", "Decapado",
    0.059, 0.060, 0.059, 0.059, 0.059, 0.060, 0.059, 0.059, 0.059, 0.060, 0.059, 0.059,
    45, 2029,
    "COL-2603210220", "LOT-23587",
    "Excelente estado superficial"
]
ws.append(example_row_dec)

# Agregar validaciones de listas desplegables
from openpyxl.worksheet.datavalidation import DataValidation

# Validación para Calibre (B3:B100)
dv_calibre = DataValidation(type="list", formula1='"CAL 10,CAL 12,CAL 14,CAL 16"', allow_blank=True)
dv_calibre.error = 'El calibre ingresado debe pertenecer a la lista del catálogo (CAL 10, CAL 12, CAL 14, CAL 16).'
dv_calibre.errorTitle = 'Calibre Inválido'
dv_calibre.prompt = 'Seleccione el calibre del catálogo'
dv_calibre.promptTitle = 'Calibre'
ws.add_data_validation(dv_calibre)
dv_calibre.add("B3:B100")

# Validación para Galvanizado_o_Decapado (C3:C100)
dv_tipo = DataValidation(type="list", formula1='"Galvanizado,Decapado,Aluminio"', allow_blank=True)
dv_tipo.error = 'El tipo de material debe ser Galvanizado, Decapado o Aluminio.'
dv_tipo.errorTitle = 'Material Inválido'
dv_tipo.prompt = 'Seleccione el tipo de material'
dv_tipo.promptTitle = 'Material'
ws.add_data_validation(dv_tipo)
dv_tipo.add("C3:C100")

# Estilos de openpyxl
fill_header = PatternFill(start_color="D32F2F", end_color="D32F2F", fill_type="solid")
font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
font_example = Font(name="Calibri", size=11, italic=True, color="555555")

align_center = Alignment(horizontal="center", vertical="center")
align_left = Alignment(horizontal="left", vertical="center")

border_thin = Side(border_style="thin", color="D3D3D3")
border_cell = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)

ws.row_dimensions[1].height = 24
ws.row_dimensions[2].height = 24

# Formatear Cabeceras (Fila 1 y Fila 2)
for r in [1, 2]:
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=r, column=col_idx)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = align_center
        cell.border = border_cell

# Formatear Filas de ejemplo (Fila 3 y Fila 4)
for row_idx in [3, 4]:
    ws.row_dimensions[row_idx].height = 20
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.font = font_example
        cell.border = border_cell
        if col_idx < len(headers):
            cell.alignment = align_center
        else:
            cell.alignment = align_left
            
for col in ws.columns:
    col_idx = col[0].column
    col_letter = get_column_letter(col_idx)
    if 4 <= col_idx <= 15:
        ws.column_dimensions[col_letter].width = 8.5
    else:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)
    
wb.save(plantilla_path)
print("Plantilla Excel de operador recreada exitosamente con 20 columnas, doble cabecera y listas desplegables.")

print("Inicialización completa.")
