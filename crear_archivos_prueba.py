import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = r"C:\Users\albertol\.gemini\antigravity\scratch\incoming_calidad"
TEST_DIR = os.path.join(BASE_DIR, "formatos_prueba")
os.makedirs(TEST_DIR, exist_ok=True)

headers = [
    "No_Atado", "Calibre", "Galvanizado_o_Decapado", 
    "Espesor_Medido_1_in", "Espesor_Medido_2_in", "Espesor_Medido_3_in", 
    "Cantidad_Hojas", "Peso_Total_Kg", 
    "Num_Colada", "Lote_Heat", 
    "Observaciones"
]

def crear_archivo_excel(filename, rows_data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Incoming_Mediciones"
    ws.append(headers)
    
    for row in rows_data:
        ws.append(row)
        
    fill_header = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_cell = Font(name="Calibri", size=11)
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    
    border_thin = Side(border_style="thin", color="D3D3D3")
    border_cell = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    
    ws.row_dimensions[1].height = 28
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = align_center
        cell.border = border_cell
        
    for r_idx in range(2, len(rows_data) + 2):
        ws.row_dimensions[r_idx].height = 20
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=r_idx, column=col_idx)
            cell.font = font_cell
            cell.border = border_cell
            if col_idx in [1, 2, 3, 7, 8, 9, 10]:
                cell.alignment = align_center
            elif col_idx in [4, 5, 6]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left
                
    # Agregar validaciones de listas desplegables
    from openpyxl.worksheet.datavalidation import DataValidation
    dv_calibre = DataValidation(type="list", formula1='"CAL 10,CAL 12,CAL 14,CAL 16"', allow_blank=True)
    ws.add_data_validation(dv_calibre)
    dv_calibre.add("B2:B100")

    dv_tipo = DataValidation(type="list", formula1='"Galvanizado,Decapado,Aluminio"', allow_blank=True)
    ws.add_data_validation(dv_tipo)
    dv_tipo.add("C2:C100")

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)
        
    file_path = os.path.join(TEST_DIR, filename)
    wb.save(file_path)
    print(f"Archivo de prueba creado: {file_path}")

# Lote 1: Aceptado - Galvanizado Calibre 14
lote1_data = [
    [
        "AT-T-GALV-14-01", "CAL 14", "Galvanizado",
        0.0785, 0.0792, 0.0788,
        150, 2450,
        "COL-1111", "HEAT-111",
        "Material aceptado, en tolerancia de espesor."
    ],
    [
        "AT-T-GALV-14-02", "CAL 14", "Galvanizado",
        0.0795, 0.0789, 0.0791,
        148, 2420,
        "COL-1112", "HEAT-111",
        "Excelente estado."
    ]
]
crear_archivo_excel("Plantilla_Prueba_Lote_1_Aceptado_GALV.xlsx", lote1_data)

# Lote 2: Aceptado - Decapado Calibre 12
lote2_data = [
    [
        "AT-N-DECP-12-01", "CAL 12", "Decapado",
        0.1042, 0.1048, 0.1045,
        110, 2350,
        "COL-2221", "HEAT-222",
        "Lámina decapada aceitada correcta."
    ],
    [
        "AT-N-DECP-12-02", "CAL 12", "Decapado",
        0.1052, 0.1049, 0.1051,
        112, 2390,
        "COL-2222", "HEAT-222",
        "Decapada sin problemas."
    ]
]
crear_archivo_excel("Plantilla_Prueba_Lote_2_Aceptado_DECP.xlsx", lote2_data)

# Lote 3: Rechazado - Galvanizado Calibre 10 (Espesores fuera de tolerancias de 0.138 +/- 0.008 -> max 0.146)
lote3_data = [
    [
        "AT-T-GALV-10-01", "CAL 10", "Galvanizado",
        0.1485, 0.1492, 0.1480, # Excede espesor max de 0.146 in
        90, 2400,
        "COL-3331", "HEAT-333",
        "Fuera de tolerancia de espesor superior al límite."
    ],
    [
        "AT-T-GALV-10-02", "CAL 10", "Galvanizado",
        0.1385, 0.1378, 0.1390,
        80, 2480,
        "COL-3332", "HEAT-333",
        "Atado de control correcto."
    ]
]
crear_archivo_excel("Plantilla_Prueba_Lote_3_Rechazado_GALV.xlsx", lote3_data)

print("--- TODOS LOS ARCHIVOS DE PRUEBA GENERADOS ---")
